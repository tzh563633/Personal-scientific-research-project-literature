import asyncio
import io
import zipfile
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from backend.app import models
from backend.app.config import settings
from backend.app.services import backup as backup_service
from backend.app.services import excel as excel_service
from backend.app.services import reviews as reviews_service
from backend.app.services.llm import MockLLMProvider
from backend.app.services import papers as papers_service
from backend.app.services.crypto import decrypt_secret, encrypt_secret
from backend.app.services.files import safe_extract_archive
from backend.app.services.folders import validate_host_folder_path, validate_relative_path
from backend.app.services.network import validate_public_url
from backend.app.services.papers import _parse_citations, _parse_references


def test_mock_provider_is_deterministic():
    provider = MockLLMProvider()
    first = provider.extract_metadata("A title\nAbstract: useful text\n2024", "paper.pdf")
    second = provider.extract_metadata("A title\nAbstract: useful text\n2024", "paper.pdf")
    assert first == second
    assert first["year"] == 2024


def test_mock_provider_skips_pdf_cover_metadata():
    provider = MockLLMProvider()
    result = provider.extract_metadata(
        "中图分类号：\nX82\n论文编号：\n成长型矿业城市韧性综合评价研究\n作者姓名：\n代大为\n2025",
        "paper.pdf",
    )
    assert result["title"] == "成长型矿业城市韧性综合评价研究"
    result = provider.extract_metadata(
        "地理研究\n第44卷 第2期\n城市何以更加“韧性”\n——数字经济的赋能效应\n于斌斌\n摘要：数字经济提升城市韧性。",
        "paper.pdf",
    )
    assert result["title"] == "城市何以更加“韧性”——数字经济的赋能效应"


def test_reference_and_citation_parsing():
    text = "A claim [1]. Another claim [2].\n\nReferences\n[1] Author. First.\n[2] Author. Second."
    assert _parse_references(text) == ["Author. First.", "Author. Second."]
    assert [item[0] for item in _parse_citations(text)] == [1, 2]


def test_archive_path_traversal_is_rejected(tmp_path: Path):
    archive = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr("../escape.txt", "bad")
    try:
        safe_extract_archive(archive, tmp_path / "out")
    except ValueError as exc:
        assert "path traversal" in str(exc)
    else:
        raise AssertionError("unsafe archive was accepted")


def test_archive_backslash_path_traversal_is_rejected(tmp_path: Path):
    archive = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr(r"..\escape.txt", "bad")
    try:
        safe_extract_archive(archive, tmp_path / "out")
    except ValueError as exc:
        assert "path traversal" in str(exc)
    else:
        raise AssertionError("unsafe backslash path was accepted")


def test_pdf_upload_checks_signature(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))

    async def save():
        upload = UploadFile(
            filename="paper.pdf",
            file=io.BytesIO(b"not a pdf"),
            headers={"content-type": "application/pdf"},
        )
        from backend.app.services.files import save_upload

        return await save_upload(upload, "uploads", {".pdf"})

    try:
        asyncio.run(save())
    except ValueError as exc:
        assert "signature" in str(exc)
    else:
        raise AssertionError("invalid PDF signature was accepted")


def test_doi_extraction_strips_trailing_punctuation():
    assert (
        papers_service._extract_doi("See https://doi.org/10.1234/example.test.")
        == "10.1234/example.test"
    )


def test_pdf_mojibake_is_repaired():
    assert papers_service._repair_mojibake("ÖÐÍ¼·ÖÀàºÅ") == "中图分类号"


def test_grobid_metadata_is_parsed(monkeypatch, tmp_path: Path):
    class Response:
        text = """
        <TEI xmlns="http://www.tei-c.org/ns/1.0">
          <teiHeader>
            <fileDesc>
              <titleStmt><title>Parsed title</title></titleStmt>
            </fileDesc>
            <profileDesc><abstract>Parsed abstract</abstract></profileDesc>
          </teiHeader>
          <text><back><listBibl><biblStruct>Reference one</biblStruct></listBibl></back></text>
        </TEI>
        """

        def raise_for_status(self):
            return None

    monkeypatch.setattr(papers_service.httpx, "post", lambda *args, **kwargs: Response())
    monkeypatch.setattr(settings, "grobid_url", "http://grobid:8070")
    source = tmp_path / "paper.pdf"
    source.write_bytes(b"%PDF-1.7\n")
    assert papers_service._try_grobid(source) == {
        "title": "Parsed title",
        "abstract": "Parsed abstract",
        "references": ["Reference one"],
    }


def test_encrypted_config_round_trip():
    encrypted = encrypt_secret("private-value")
    assert encrypted != "private-value"
    assert decrypt_secret(encrypted) == "private-value"


def test_private_source_urls_require_explicit_allowlist(monkeypatch):
    monkeypatch.setattr(
        "backend.app.services.network.socket.getaddrinfo",
        lambda *args, **kwargs: [(None, None, None, None, ("127.0.0.1", 80))],
    )
    monkeypatch.setattr(settings, "outbound_allowed_hosts", "")
    try:
        validate_public_url("http://host.docker.internal:8765/feed.xml")
    except ValueError as exc:
        assert "private" in str(exc)
    else:
        raise AssertionError("private source URL was accepted without an allowlist")

    monkeypatch.setattr(settings, "outbound_allowed_hosts", "host.docker.internal")
    validate_public_url("http://host.docker.internal:8765/feed.xml")


def test_backup_storage_excludes_backup_directory(tmp_path: Path):
    storage = tmp_path / "storage"
    backup_root = storage / "backups"
    backup_root.mkdir(parents=True)
    (storage / "normal.txt").write_text("keep", encoding="utf-8")
    (backup_root / "old.zip").write_bytes(b"ignore")
    archive = tmp_path / "storage.zip"
    backup_service._archive_storage(storage, backup_root, archive)
    with zipfile.ZipFile(archive) as handle:
        assert handle.namelist() == ["normal.txt"]


def test_excel_preserves_manual_visible_edits(tmp_path: Path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", connect_args={"check_same_thread": False})
    models.Base.metadata.create_all(engine)
    workbook_path = tmp_path / "papers.xlsx"
    monkeypatch.setattr(excel_service, "EXCEL_PATH", workbook_path)

    with Session(engine) as db:
        paper = models.Paper(title="Original title", authors="Author", year=2024)
        db.add(paper)
        db.commit()
        excel_service.generate_excel(db)

        workbook = load_workbook(workbook_path)
        sheet = workbook["Papers"]
        sheet.cell(2, 2).value = "Human-edited title"
        workbook.save(workbook_path)

        paper.title = "Platform title changed"
        db.commit()
        excel_service.generate_excel(db)

        workbook = load_workbook(workbook_path)
        assert workbook["Papers"].cell(2, 2).value == "Human-edited title"
        assert db.query(models.ManualEdit).count() == 1


def test_folder_path_validation_requires_absolute_paths():
    assert validate_host_folder_path(r"D:\research\papers") == r"D:\research\papers"
    assert validate_relative_path(r"nested\paper.pdf") == "nested/paper.pdf"
    for value in ("relative/path", "", "../paper.pdf"):
        try:
            if value.endswith(".pdf"):
                validate_relative_path(value)
            else:
                validate_host_folder_path(value)
        except ValueError:
            pass
        else:
            raise AssertionError(f"unsafe folder path accepted: {value}")


def test_review_generation_uses_selected_excel_and_records_traceability(tmp_path: Path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 'review.db'}", connect_args={"check_same_thread": False})
    models.Base.metadata.create_all(engine)
    storage_root = tmp_path / "storage"
    monkeypatch.setattr(settings, "storage_root", str(storage_root))
    workbook_path = storage_root / "exports" / "selected.xlsx"
    workbook_path.parent.mkdir(parents=True)
    from openpyxl import Workbook

    sheet_book = Workbook()
    sheet = sheet_book.active
    sheet.title = "Papers"
    sheet.append(["paper_id", "标题"])
    sheet.append([1, "Excel-guided resilience paper"])
    sheet_book.save(workbook_path)

    class Provider:
        def generate_review(self, framework, sources):
            assert sources[0]["source"] == "excel"
            return "# Excel Review"

    monkeypatch.setattr(reviews_service, "search_academic_sources", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(reviews_service, "get_provider", lambda *_args, **_kwargs: Provider())

    with Session(engine) as db:
        paper = models.Paper(
            id=1,
            title="Excel-guided resilience paper",
            status="processed",
            citation_gbt="Author. Excel-guided resilience paper.",
        )
        framework = models.ReviewFramework(
            name="Excel review",
            content="resilience review outline",
            excel_path="exports/selected.xlsx",
        )
        db.add_all([paper, framework])
        db.commit()
        db.refresh(framework)

        output = reviews_service.generate_review(db, framework)

        assert output.content == "# Excel Review"
        assert output.source_count == 1
        assert output.verified_source_count == 1
        assert output.full_text_source_count == 1
        assert output.fact_check_summary["passed"] == 1
        source = db.query(models.ReviewSource).one()
        assert source.source_type == "excel"
        assert source.verified is True
        assert (storage_root / output.missing_pdf_md_path).is_file()


def test_transient_deepseek_key_is_used_without_persisting_it(tmp_path: Path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 'transient-review.db'}", connect_args={"check_same_thread": False})
    models.Base.metadata.create_all(engine)
    monkeypatch.setattr(settings, "storage_root", str(tmp_path / "storage"))
    captured = {}

    class Provider:
        def generate_review(self, _framework, _sources):
            return "# Transient Key Review"

    def provider_factory(provider_name=None, api_key=None):
        captured["provider_name"] = provider_name
        captured["api_key"] = api_key
        return Provider()

    monkeypatch.setattr(reviews_service, "search_academic_sources", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(reviews_service, "get_provider", provider_factory)

    with Session(engine) as db:
        db.add(models.Paper(title="Digital resilience", status="processed"))
        framework = models.ReviewFramework(name="Transient", content="digital resilience")
        db.add(framework)
        db.commit()
        db.refresh(framework)

        output = reviews_service.generate_review(
            db,
            framework,
            transient_deepseek_api_key="temporary-deepseek-key",
        )

        assert output.content == "# Transient Key Review"
        assert captured == {"provider_name": "deepseek", "api_key": "temporary-deepseek-key"}
        assert db.query(models.SystemConfig).count() == 0
