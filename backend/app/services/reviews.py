from __future__ import annotations

import re
from pathlib import Path
from pathlib import PurePosixPath

from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..config import settings
from ..models import Job, Paper, PaperFile, ReviewFramework, ReviewOutput, ReviewSource, SystemConfig
from .academic import (
    _normalize_doi,
    download_source_pdf,
    fact_check,
    search_academic_sources,
    verify_external_source,
)
from .files import absolute_storage_path, file_metadata, relative_storage_path, sha256_path
from .llm import get_provider
from .papers import process_paper
from .crypto import decrypt_secret


def _keywords(content: str) -> list[str]:
    tokens = re.findall(r"[\w\u4e00-\u9fff]{2,}", content.lower())
    stop_words = {"研究", "分析", "综述", "以及", "相关", "问题", "方法", "the", "and", "of"}
    return [token for token in dict.fromkeys(tokens) if token not in stop_words][:20]


def _local_papers(db: Session, keywords: list[str]) -> list[Paper]:
    if not keywords:
        return []
    filters = []
    for keyword in keywords:
        pattern = f"%{keyword}%"
        filters.extend(
            [
                Paper.title.ilike(pattern),
                Paper.abstract.ilike(pattern),
                Paper.core_topics.ilike(pattern),
                Paper.secondary_topics.ilike(pattern),
                Paper.innovation_points.ilike(pattern),
            ]
        )
    found = {paper.id: paper for paper in db.query(Paper).filter(or_(*filters)).limit(100).all()}
    for paper in db.query(Paper).filter(Paper.status == "processed").limit(200).all():
        text_files = db.query(PaperFile).filter(PaperFile.paper_id == paper.id, PaperFile.kind == "txt").all()
        for file in text_files:
            try:
                text = absolute_storage_path(file.path).read_text(encoding="utf-8", errors="ignore").lower()
            except (OSError, ValueError):
                continue
            if any(keyword in text for keyword in keywords):
                found[paper.id] = paper
                break
    return list(found.values())[:100]


def _resolve_excel_export(excel_path: str) -> Path:
    normalized = (excel_path or "").strip().replace("\\", "/")
    relative = PurePosixPath(normalized)
    if (
        not normalized
        or relative.is_absolute()
        or any(part in {"", ".", ".."} for part in relative.parts)
        or not normalized.startswith("exports/")
        or relative.suffix.lower() not in {".xlsx", ".xlsm"}
    ):
        raise ValueError("Review Excel path must be a relative exports/*.xlsx or exports/*.xlsm path")
    export_root = (settings.storage_path / "exports").resolve()
    resolved = (settings.storage_path / relative).resolve()
    if export_root not in resolved.parents or not resolved.is_file():
        raise ValueError("Selected review Excel file does not exist in platform exports")
    return resolved


def _papers_from_excel(db: Session, excel_path: str | None) -> list[Paper]:
    if not excel_path:
        return []
    workbook_path = _resolve_excel_export(excel_path)
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["Papers"] if "Papers" in workbook.sheetnames else workbook.active
        headers = [str(value or "").strip().lower() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        paper_column = headers.index("paper_id") if "paper_id" in headers else 0
        papers: list[Paper] = []
        seen: set[int] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_id = row[paper_column] if paper_column < len(row) else None
            try:
                paper_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if paper_id in seen:
                continue
            paper = db.get(Paper, paper_id)
            if paper:
                seen.add(paper_id)
                papers.append(paper)
        return papers
    finally:
        try:
            workbook.close()
        except (UnboundLocalError, AttributeError):
            pass


def _saved_deepseek_key(db: Session) -> str | None:
    item = db.query(SystemConfig).filter(SystemConfig.key == "DEEPSEEK_API_KEY").first()
    if not item or not item.value:
        return None
    try:
        return decrypt_secret(item.value)
    except Exception:
        return None


def _review_provider(db: Session, transient_deepseek_api_key: str | None):
    if transient_deepseek_api_key:
        return get_provider("deepseek", transient_deepseek_api_key)
    saved_key = _saved_deepseek_key(db)
    if saved_key:
        return get_provider("deepseek", saved_key)
    return get_provider()


def _ingest_external_pdf(db: Session, source: dict, path: Path) -> Paper | None:
    digest = sha256_path(path)
    existing_file = db.query(PaperFile).filter(PaperFile.sha256 == digest).first()
    if existing_file:
        return db.get(Paper, existing_file.paper_id)
    metadata = file_metadata(path.name, "application/pdf")
    paper = Paper(
        title=source.get("title") or path.stem,
        authors=source.get("authors"),
        year=source.get("year"),
        doi=_normalize_doi(source.get("doi")) or None,
        abstract=source.get("abstract"),
        file_path=relative_storage_path(path),
        status="pending",
    )
    db.add(paper)
    db.flush()
    db.add(
        PaperFile(
            paper_id=paper.id,
            kind="original",
            path=relative_storage_path(path),
            sha256=digest,
            size_bytes=path.stat().st_size,
            original_name=metadata["original_name"],
            extension=metadata["extension"],
            mime_type=metadata["mime_type"],
        )
    )
    job = Job(kind="paper_processing", entity_id=paper.id, status="pending", message="External PDF queued")
    db.add(job)
    db.commit()
    db.refresh(job)
    process_paper(db, job.id)
    if job.entity_id:
        return db.get(Paper, job.entity_id)
    return None


def generate_review(
    db: Session,
    framework: ReviewFramework,
    transient_deepseek_api_key: str | None = None,
) -> ReviewOutput:
    keywords = _keywords(framework.content)
    excel_papers = _papers_from_excel(db, framework.excel_path)
    local_by_id = {paper.id: paper for paper in excel_papers}
    for paper in _local_papers(db, keywords):
        local_by_id.setdefault(paper.id, paper)
    local = list(local_by_id.values())[:100]
    sources: list[dict] = [
        {
            "title": paper.title,
            "authors": paper.authors,
            "year": paper.year,
            "doi": paper.doi,
            "citation": paper.citation_gbt,
            "source": "excel" if paper.id in {item.id for item in excel_papers} else "local",
            "verified": True,
            "verified_by": "platform_excel" if paper.id in {item.id for item in excel_papers} else "local_paper",
            "full_text_available": True,
        }
        for paper in local
    ]
    missing: list[dict] = []
    external: list[dict] = []
    for keyword in keywords[:5]:
        results = search_academic_sources(keyword, limit=3)
        external.extend(results)
    seen = set()
    for source in external:
        key = _normalize_doi(source.get("doi")) or (source.get("title") or "").lower()
        if not key or key in seen:
            continue
        seen.add(key)
        source["verified"] = verify_external_source(source)
        source["verified_by"] = "crossref" if source["verified"] else None
        source["full_text_available"] = False
        if source["verified"] and settings.supplement_download_enabled:
            downloaded = download_source_pdf(source, settings.storage_path / "uploads")
            if downloaded:
                paper = _ingest_external_pdf(db, source, downloaded)
                if paper:
                    source["full_text_available"] = True
                    source["paper_id"] = paper.id
                    source["verified_by"] = "downloaded_pdf"
        if source["verified"]:
            sources.append(source)
        if not source["verified"] or not source["full_text_available"]:
            missing.append(
                {
                    "title": source.get("title"),
                    "doi": source.get("doi"),
                    "url": source.get("url"),
                    "reason": "未核实来源" if not source["verified"] else "未下载全文",
                }
            )

    verified_sources = [source for source in sources if source.get("verified")]
    for source in verified_sources:
        source["fact_checked"] = fact_check(source, local)
    try:
        content = _review_provider(db, transient_deepseek_api_key).generate_review(
            framework.content,
            verified_sources,
        )
        model_fallback = False
    except Exception:
        content = get_provider("mock").generate_review(framework.content, verified_sources)
        content += "\n\n> 真实模型调用失败，已降级为 Mock Provider。"
        model_fallback = True
    missing_path = settings.storage_path / "reviews" / f"missing-{framework.id}.md"
    missing_lines = [
        f"- {item.get('title') or '未命名文献'}"
        f" | DOI: {item.get('doi') or '无'}"
        f" | {item.get('reason') or item.get('error') or ''}"
        f" | {item.get('url') or ''}"
        for item in missing
    ]
    missing_path.write_text(
        "# 缺失文献提醒\n\n" + "\n".join(missing_lines),
        encoding="utf-8",
    )
    output = ReviewOutput(
        framework_id=framework.id,
        content=content,
        missing_pdf_md_path=str(missing_path.relative_to(settings.storage_path)),
        source_count=len(sources),
        verified_source_count=len(verified_sources),
        full_text_source_count=sum(bool(source.get("full_text_available")) for source in verified_sources),
        fact_check_summary={
            "checked": len(verified_sources),
            "passed": sum(bool(source.get("fact_checked")) for source in verified_sources),
            "failed": sum(not bool(source.get("fact_checked")) for source in verified_sources),
            "model_fallback": model_fallback,
        },
    )
    db.add(output)
    db.flush()
    for source in verified_sources:
        db.add(
            ReviewSource(
                output_id=output.id,
                source_type=source.get("source") or "unknown",
                title=source.get("title") or "Untitled",
                authors=source.get("authors"),
                year=source.get("year"),
                doi=source.get("doi"),
                url=source.get("url"),
                verified=bool(source.get("verified")),
                full_text_available=bool(source.get("full_text_available")),
                source_metadata={
                    key: value
                    for key, value in source.items()
                    if key not in {"title", "authors", "year", "doi", "url", "verified", "full_text_available"}
                },
            )
        )
    db.commit()
    db.refresh(output)
    return output
