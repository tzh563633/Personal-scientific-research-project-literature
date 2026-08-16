from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from ..config import settings
from ..models import ExcelUpdate, ManualEdit, Paper, now

FIELDS = [
    ("title", "标题"),
    ("authors", "作者"),
    ("year", "年份"),
    ("core_topics", "核心主题"),
    ("secondary_topics", "次要主题"),
    ("abstract", "摘要"),
    ("citation_gbt", "引用格式"),
    ("innovation_points", "创新点"),
]
EXCEL_PATH = settings.storage_path / "exports" / "papers.xlsx"
EXCEL_CELL_LIMIT = 32000


def _clean_value(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))


def _display_value(value) -> str:
    text = _clean_value(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return text[:EXCEL_CELL_LIMIT] + "\n[内容过长，已截断；完整内容请查看平台文献文件]"


def _source_value(paper: Paper, field: str) -> str:
    return _clean_value(getattr(paper, field, None))


def _value(paper: Paper, field: str):
    return _display_value(getattr(paper, field, None))


def generate_excel(db: Session) -> ExcelUpdate:
    update = ExcelUpdate(status="running", update_time=now())
    db.add(update)
    db.commit()
    try:
        existing = load_workbook(EXCEL_PATH) if EXCEL_PATH.exists() else None
        old_sheet = existing["Papers"] if existing and "Papers" in existing.sheetnames else None
        old_meta = existing["_meta"] if existing and "_meta" in existing.sheetnames else None
        old_values: dict[tuple[int, str], str] = {}
        if old_sheet and old_meta:
            for row in old_meta.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    old_values[(int(row[0]), str(row[1]))] = "" if row[2] is None else str(row[2])
        old_rows: dict[int, int] = {}
        if old_sheet:
            for row_number in range(2, old_sheet.max_row + 1):
                paper_id = old_sheet.cell(row_number, 1).value
                if paper_id is not None:
                    old_rows[int(paper_id)] = row_number

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Papers"
        sheet.append(["paper_id", *[label for _, label in FIELDS], "人工状态"])
        sheet.column_dimensions["A"].hidden = True
        meta = workbook.create_sheet("_meta")
        meta.sheet_state = "hidden"
        meta.append(["paper_id", "field", "source_value", "display_value"])
        changed_ids: set[int] = set()
        for paper in db.query(Paper).order_by(Paper.id).all():
            values = []
            manual = False
            old_row = old_rows.get(paper.id)
            for column, (field, _) in enumerate(FIELDS, 2):
                source = _source_value(paper, field)
                current = _display_value(source)
                previous = old_values.get((paper.id, field))
                previous_visible = None
                if old_sheet and old_row:
                    previous_visible = old_sheet.cell(old_row, column).value
                    previous_visible = "" if previous_visible is None else str(previous_visible)
                previous_display = _display_value(previous) if previous is not None else None
                if previous is not None and previous_visible != previous_display:
                    manual = True
                    current = previous_visible
                    existing_edit = (
                        db.query(ManualEdit)
                        .filter(ManualEdit.paper_id == paper.id, ManualEdit.field_name == field)
                        .order_by(ManualEdit.edited_at.desc())
                        .first()
                    )
                    if not existing_edit or existing_edit.edited_value != current:
                        db.add(
                            ManualEdit(
                                paper_id=paper.id,
                                field_name=field,
                                original_value=previous,
                                edited_value=current,
                            )
                        )
                meta.append([paper.id, field, source, _display_value(source)])
                values.append(current)
            if manual:
                changed_ids.add(paper.id)
            sheet.append([paper.id, *values, "人工修正" if manual else "自动"])
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        workbook.save(EXCEL_PATH)
        update.status = "succeeded"
        update.added_count = len(changed_ids)
        update.preserved_manual_count = len(changed_ids)
        update.paper_count = db.query(Paper).count()
        update.error_message = None
        db.commit()
        return update
    except Exception as exc:
        db.rollback()
        update.status = "failed"
        update.error_message = str(exc)
        db.add(update)
        db.commit()
        return update
